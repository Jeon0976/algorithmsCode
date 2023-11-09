import requests
import openpyxl

from api import KAKAO_API_KEY
from config import KakaoMapAddressToCoordinate as KMAC

URL = "https://dapi.kakao.com/v2/local/search/address"

# x, y 좌표 반환 함수 
def get_coordinates(address):
    headers = { 
        "Authorization": KAKAO_API_KEY
    }

    params = { 
        "query": address
    }

    respone = requests.get(URL, headers=headers, params=params)
    
    data = respone.json()

    if not data['documents']: 
        return None,None
    return data['documents'][0]['x'], data['documents'][0]['y']

def get_coordinates_from_excel():
    wb = openpyxl.load_workbook(KMAC.EXCEL_FILENAME)
    ws = wb.worksheets[KMAC.WORKSHEET]

    row = KMAC.START_ROW
    while ws[f"{KMAC.ADDRESS_COLUMN}{row}"].value:
        address = ws[f"{KMAC.ADDRESS_COLUMN}{row}"].value
        lat, lng = get_coordinates(address)
        if lat and lng:
            ws[f"{KMAC.LATITUDE_COLUMN}{row}"].value = lat
            ws[f"{KMAC.LONGITUDE_COLUMN}{row}"].value = lng
        row += 1

    wb.save("update_" + KMAC.EXCEL_FILENAME)
    print("End Work")

if __name__ == "__main__":
    get_coordinates_from_excel()